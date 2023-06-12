############################
#Borrar pantalla

# GRUPO 18
import os
import string

def clear_screen():
    os.system('clear' if os.name == 'posix' else 'cls')
############################
#####################################################################################################
# Funciones punto 1

def calcular_indices(recuperados, relevantes_consulta,cant_reg_relevantes):
    recall = relevantes_consulta / cant_reg_relevantes # Total de documentos relevantes en la base de datos
    recall=recall*100
    recall="{:.2f}%".format(recall)
    precision = relevantes_consulta / recuperados
    precision=precision*100
    precision="{:.2f}%".format(precision)
    return recall, precision


# FIN Funciones punto 1
#####################################################################################################

#####################################################################################################
# Funciones punto 2
#####################################################################################################

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

def compare_lists(list1, list2):
    # Convertir las listas en cadenas de texto
    text1 = " ".join(list1)
    text2 = " ".join(list2)

    # Vectorización de los textos utilizando TF-IDF 
    vectorizer = TfidfVectorizer()
    tfidf = vectorizer.fit_transform([text1, text2])

    # Cálculo de la similitud coseno entre los vectores TF-IDF de los textos
    sim_cos = cosine_similarity(tfidf[0], tfidf[1])

    return sim_cos[0][0]



from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from nltk.corpus import stopwords

def compara_texto_original(text1, text2):
    # Obtener las stopwords para español y convertirlas a lista
    stop_words = list(set(stopwords.words('spanish')))
    
    # Vectorización de los textos utilizando TF-IDF sin eliminar stopwords
    vectorizer = TfidfVectorizer(stop_words=stop_words)
    tfidf = vectorizer.fit_transform([text1, text2])

    # Cálculo de la similitud coseno entre los vectores TF-IDF de los textos
    sim_cos = cosine_similarity(tfidf[0], tfidf[1])

    return sim_cos[0][0]

    


################################################################
import fitz
def leer_pdf(nombre_archivo):
    texto_completo = ""
    
    with fitz.open(nombre_archivo) as doc:
        for pagina in doc:
            texto_completo += pagina.get_text()
    
    return texto_completo

################################################################

def leer_txt(nombre_archivo):
    texto_completo = ""

    with open(nombre_archivo, "r", encoding="utf-8") as archivo:
        texto_completo = archivo.read()

    return texto_completo

################################################################

from nltk.corpus import stopwords

def eliminar_stopwords(lista_palabras):
    # Obtenemos las stopwords para español
    stop_words_es = set(stopwords.words('spanish'))
    # Obtenemos las stopwords para inglés
    stop_words_en = set(stopwords.words('english'))

    stop_words_personalizadas = ['cada','este','cómo']
    # Combinamos ambos conjuntos de stopwords
    stop_words = stop_words_es.union(stop_words_en).union(stop_words_personalizadas)

    # Eliminamos las stopwords de la lista de palabras
    lista_sin_stopwords = [palabra for palabra in lista_palabras if not palabra.lower() in stop_words]
    
    return lista_sin_stopwords

#####################################################################
import string

def eliminar_caracteres(cadena):
    caracteres = ['"', "'", ",", ";",":",".","-","“","”","°"]
    for caracter in caracteres:
        cadena = cadena.replace(caracter, "")
    return cadena

#####################################################################
import re

def eliminar_links(cadena):
    # Expresión regular para detectar enlaces
    patron = r"http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+"
    
    # Eliminar los enlaces utilizando la expresión regular
    cadena_sin_links = re.sub(patron, "", cadena)
    return cadena_sin_links



def eliminar_parentesis_corchetes(lista):
    lista_sin_parentesis_corchetes = []
    for elemento in lista:
        elemento_sin_parentesis_corchetes = elemento.replace("(", "").replace(")", "").replace("[", "").replace("]", "")
        lista_sin_parentesis_corchetes.append(elemento_sin_parentesis_corchetes)
    return lista_sin_parentesis_corchetes

import re


def eliminar_numeros_lista(lista):
    patron_numeros = r'\d+(?:,\d+)*(?:\.\d+)?'
    lista_sin_numeros = [re.sub(patron_numeros, '', cadena).strip() for cadena in lista]
    return lista_sin_numeros


def eliminar_caracteres_unicos(lista):
    lista_filtrada = [palabra for palabra in lista if len(set(palabra)) > 1]
    return lista_filtrada


#####################################################################
from nltk.stem import SnowballStemmer

def algoritmo_snowball(lista_palabras):
    # Se crea una instancia del stemmer de Snowball para español
    stemmer = SnowballStemmer("spanish")
    # Aplicamos stemming a cada palabra
    stemmed = [stemmer.stem(palabra) for palabra in lista_palabras]
    
    return stemmed

def obtener_bigramas(lista_elementos):
    # Obtenemos los bigramas
    bigramas = [lista_elementos[i] + " " + lista_elementos[i+1] for i in range(len(lista_elementos)-1)]
    return bigramas


# FIN Funciones punto 2
#####################################################################################################
import openpyxl

def generar_excel(ranking,nombre,color):

    # Crear un nuevo archivo Excel
    libro_excel = openpyxl.Workbook()
    hoja = libro_excel.active

    # Agregar encabezados
    hoja.cell(row=1, column=1).value = "Ranking"
    hoja.cell(row=1, column=2).value = "Score"
    hoja.cell(row=1, column=3).value = "Documento"
    hoja.cell(row=1, column=4).value = "Contenido"

    # Agregar color a las celdas de los títulos de noticias #ELiminar si esta mal
    for i in range(1, 5):
        hoja.cell(row=1, column=i).fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')

    for i in range(2, 5): 
        hoja.cell(row=i, column=1).fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
   

    # Guardar los resultados en el archivo Excel
    fila = 2
    for i, (similitud, documento, contenido) in enumerate(ranking):
        hoja.cell(row=fila, column=1).value = i+1
        hoja.cell(row=fila, column=2).value="{:.3f}".format(similitud)
        hoja.cell(row=fila, column=3).value = documento
        hoja.cell(row=fila, column=4).value = contenido
        fila += 1

    # Guardar el archivo Excel
    libro_excel.save(nombre)


from openpyxl import load_workbook

def modificar_formato_columnas_xlsx(nom_archivo):
    # Cargar el archivo existente
    book = load_workbook(nom_archivo)

    # Seleccionar la hoja a modificar
    sheet = book.active

    # Modificar el formato de las columnas
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    # Guardar los cambios en el archivo existente
    book.save(nom_archivo)

####################################################################################################

def mostrar_ranking(similitudes):
    similitudes.sort(reverse=True)
    ranking = similitudes[:3]

    print("\nLos 3 documentos con mayor similitud con el {} son:\n".format(elemento))
    for i, (similitud, documento,contenido) in enumerate(ranking):
        print("Rank {}: Similitud coseno: {:.2f} - Documento: {}".format(i+1, similitud, documento))
    return ranking


#####################################################################################################

while True:
    clear_screen()
    print("Trabajo Practico N°5 Clasificación Automática de Textos"+"\n\n")
    print("1. Calculo de la medición de índices de recuperación y precisión en diferentes búsquedas")
    print("2. Comparación de representaciones TF-IDF en la búsqueda de documentos similares en distintos casos de procesamiento")
    print("3. Salir\n")

    opcion = input("Ingrese una opción: ")
    if opcion == "1":
        clear_screen()
        """Supongamos que partimos de una base de datos con 100 registros, 
        de los cuales 40 son relevantes para nuestra consulta y los otros 60 no lo son. 
        A partir de diferentes resultados de nuestras búsquedas, medir los índices de recuperación y de precisión:"""


        cant_reg_base_dato=int(input("Ingrese la cantidad de registros de la base de datos: "))
        cant_reg_relevantes=int(input("Ingrese cantidad de ducumentos relevantes para la consulta: "))
        cant_no_relevantes=cant_reg_base_dato-cant_reg_relevantes
        print("Cantidad de documentos no relevantes para la consulta:",cant_no_relevantes)
        print("\n")
        recuperados = int(input("Ingrese el número de documentos recuperados: "))
        relevantes_consulta = int(input("Ingrese el número de documentos relevantes para la consulta: "))

        recall, precision = calcular_indices(recuperados, relevantes_consulta,cant_reg_relevantes)

        print("\nÍndice de Recuperación: ", recall)
        print("\nÍndice de Precisión: ", precision)

        input("\nPresione enter para continuar...")

           
    
    elif opcion == "2":


        #Consigue la ruta donde se encuentran almacenados los documentos PDFs y TXTs

        ruta_directorio = r"..\TPN5" # Reemplaza con la ruta de tu directorio de documentos
        lista_archivos = os.listdir(ruta_directorio)

        lista_archivos_pdf = [archivo for archivo in lista_archivos if archivo.endswith('.pdf')]
        lista_archivos_txt = [archivo for archivo in lista_archivos if archivo.endswith('.txt')]
        

        documentos_pdf = []
        for archivo in lista_archivos_pdf:
            # Leer el archivo PDF de la lista de PDFs
            texto_documentos = leer_pdf(archivo)
            documentos_pdf.append(texto_documentos)


        documentos_txt = []
                
        for archivo in lista_archivos_txt:
            # Leer el archivo PDF de la lista de TXTs
            texto_documentos = leer_txt(archivo)
            documentos_txt.append(texto_documentos)
        
        
        
        clear_screen()
        print("Se  necesita  definir el documento PDF para usar en la prueba de comparacion\n")       
        while True:
                    try:
                        numero = int(input("Ingresa un numero del documento de prueba del 1 al 5: "))
                        if 1 <= numero <= 5:
                            break
                        else:
                            print("El número debe estar en el rango del 1 al 5.")
                            
                    except ValueError:
                        print("Error: Debes ingresar un número entero.")
        

        print("\nAhora se necesita  definir el documento TXT para usar en la prueba de comparacion\n")       
        while True:
                    try:
                        numero2 = int(input("Ingresa un numero del documento de prueba del 1 al 10: "))
                        if 1 <= numero2 <= 10:
                            break
                        else:
                            print("El número debe estar en el rango del 1 al 10.")
                            
                    except ValueError:
                        print("Error: Debes ingresar un número entero.")
        
        input("\nPresione enter para continuar al submenu de opciones...")


        docupdflistlimpios=[]

        for docupdf in documentos_pdf:
            docupdf=eliminar_links(docupdf)
            docupdf=eliminar_caracteres(docupdf)

            lista = docupdf.split()
            lista = [elemento.lower() for elemento in lista]
            
            lista=eliminar_parentesis_corchetes(lista)
            lista=eliminar_numeros_lista(lista)
            lista = list(filter(None, lista))
            lista=eliminar_stopwords(lista)
            lista=eliminar_caracteres_unicos(lista)
            docupdflistlimpios.append(lista)
        
        docutxtlistlimpios=[]

        for docutxt in documentos_txt:
            docutxt=eliminar_links(docutxt)
            docupdf=eliminar_caracteres(docutxt)

            lista = docutxt.split()
            lista = [elemento.lower() for elemento in lista]
            
            lista=eliminar_parentesis_corchetes(lista)
            lista=eliminar_numeros_lista(lista)
            lista = list(filter(None, lista))
            lista=eliminar_stopwords(lista)
            lista=eliminar_caracteres_unicos(lista)
            docutxtlistlimpios.append(lista)


##########################################################################################


        #SUB MENU
        while True:
            clear_screen()
            print("¡Ahora selecciona un opcion para lo cual trabajar con el documento de prueba!"+"\n\n")
            print("1. Con el texto original")
            print("2. Eliminando stop-words")
            print("3. Realizando stemming")
            print("4. Con bi-gramas")
            print("5. Volver al menú principal\n")

            subopcion = input("Ingrese una opción: ")
            if subopcion == "1":
                clear_screen()
                #REP A

                print("REP A Texto original de los documentos PDFs")
                # Obtener un índice dentro del rango de la lista
                indice_docu = numero-1
                # Obtener el elemento correspondiente al índice aleatorio
                elemento = lista_archivos_pdf[indice_docu]
                # Imprimir el índice y el elemento elegido
                print("\nEl documento elegido para la prueba es el : {}\n".format(elemento))
                copialistapdfs=list(lista_archivos_pdf)
                # Eliminar el elemento en el índice de la lista necesario para evitar valores repetidos
                del copialistapdfs[indice_docu]
                #Copia de la lista de documentos original para trabajar
                aux_documentos_pdf=list(documentos_pdf)
                del aux_documentos_pdf[indice_docu]
                #Se establece los elementos del documento de prueba
                docprueba=documentos_pdf[indice_docu]


                # Se comienza a hacer la comparación
                max_similitud = 0.0
                doc1_max = ""
                similitudes = []
                print("\nResultado de comparaciones...\n")

                for i in range(len(aux_documentos_pdf)):
                    similitud = compara_texto_original(docprueba, aux_documentos_pdf[i])
                    if similitud > max_similitud:
                        max_similitud = similitud
                        doc1_max = copialistapdfs[i]
                    print("Similitud coseno: {:.2f} entre el {} con el {}".format(similitud, elemento, copialistapdfs[i]))
                    similitudes.append((similitud, copialistapdfs[i],aux_documentos_pdf[i]))
                print("\nLa máxima similitud de {:.2f} se encuentra entre el {} y el {}\n".format(max_similitud,elemento,doc1_max))

                ranking=mostrar_ranking(similitudes)
                
                #Se restaura los valores originales para no alterar en una segunda iteracion
                copialistapdfs=list(lista_archivos_pdf)
                aux_documentos_pdf=list(documentos_pdf)

                generar_excel(ranking,"RepA_TextoOriginal_PDFs.xlsx","FFFF00")
                modificar_formato_columnas_xlsx("RepA_TextoOriginal_PDFs.xlsx")

                print("\nSe genero un archivo Excel con los resultados del REP A...\n")
                input("\nPresione enter para continuar...\n")
                clear_screen()

                #######################################################################################################################
                # REP B
                clear_screen()
                print("REP B Texto original de los documentos TXTs")

                # Obtener un índice dentro del rango de la lista
                indice_docu = numero2-1
                # Obtener el elemento correspondiente al índice aleatorio
                elemento = lista_archivos_txt[indice_docu]
                # Imprimir el índice y el elemento elegido
                print("\nEl documento elegido para la prueba es el : {}\n".format(elemento))
                copialistatxts=list(lista_archivos_txt)
                # Eliminar el elemento en el índice de la lista necesario para evitar valores repetidos
                del copialistatxts[indice_docu]

                #Copia de la lista de documentos original para trabajar
                aux_documentos_txt=list(documentos_txt)
                del aux_documentos_txt[indice_docu]
                #Se establece los elementos del documento de prueba
                docprueba=documentos_txt[indice_docu]


                # Se comienza a hacer la comparación
                max_similitud = 0.0
                doc1_max = ""
                similitudes = []
                print("\nResultado de comparaciones...\n")

                for i in range(len(aux_documentos_txt)):
                    similitud = compara_texto_original(docprueba, aux_documentos_txt[i])
                    if similitud > max_similitud:
                        max_similitud = similitud
                        doc1_max = copialistatxts[i]
                    print("Similitud coseno: {:.2f} entre el {} con el {}".format(similitud, elemento, copialistatxts[i]))
                    similitudes.append((similitud, copialistatxts[i],aux_documentos_txt[i]))
                print("\nLa máxima similitud de {:.2f} se encuentra entre el {} y el {}\n".format(max_similitud,elemento,doc1_max))

                ranking=mostrar_ranking(similitudes)
                
                
                #Se restaura los valores originales para no alterar en una segunda iteracion
                copialistatxts=list(lista_archivos_txt)
                aux_documentos_txt=list(documentos_txt)

                generar_excel(ranking,"RepB_TextoOriginal_TXTs.xlsx","FFFF00")
                modificar_formato_columnas_xlsx("RepB_TextoOriginal_TXTs.xlsx")
                print("\nSe genero un archivo Excel con los resultados del REP B...\n")


##################################################################################
                input("\nPresione enter para continuar...")
            elif subopcion == "2":


                clear_screen()
                print("REP A Texto original sin STOPWORDS de los documentos PDFs")

                # Obtener un índice dentro del rango de la lista
                indice_docu = numero-1
                # Obtener el elemento correspondiente al índice aleatorio
                elemento = lista_archivos_pdf[indice_docu]
                # Imprimir el índice y el elemento elegido
                print("\nEl documento elegido para la prueba es el : {}\n".format(elemento))
                copialistapdfs=list(lista_archivos_pdf)
                # Eliminar el elemento en el índice de la lista necesario para evitar valores repetidos
                del copialistapdfs[indice_docu]



                #Copia de la lista de documentos sin stopwords para trabajar
                aux_documentos_pdf=list(docupdflistlimpios)

                #Se establece los elementos del documento de prueba
                docprueba=aux_documentos_pdf[indice_docu]
                del aux_documentos_pdf[indice_docu]


                #Se comienza hacer la comparacion
                max_similitud = 0.0
                doc1_max = ""
                similitudes = []

                print("\nResultado de comparaciones...\n")
                for i in range(len(aux_documentos_pdf)):
                    similitud = compare_lists(docprueba, aux_documentos_pdf[i])
                    if similitud > max_similitud:
                        max_similitud = similitud
                        doc1_max = copialistapdfs[i]
                    print("Similitud coseno: {:.2f} entre el {} con el  {}".format(similitud,elemento,copialistapdfs[i]))
                    similitudes.append((similitud, copialistapdfs[i],' '.join(aux_documentos_pdf[i])))
                print("\nLa máxima similitud de {:.2f} se encuentra entre el {} y el {}\n".format(max_similitud,elemento,doc1_max))

                
                ranking=mostrar_ranking(similitudes)

                #Se restaura los valores originales para no alterar en una segunda iteracion
                copialistapdfs=list(lista_archivos_pdf)
                aux_documentos_pdf=list(docupdflistlimpios)

                generar_excel(ranking,"RepA_SinStopWords_PDFs.xlsx","FFA500")
                modificar_formato_columnas_xlsx("RepA_SinStopWords_PDFs.xlsx")
                print("\nSe genero un archivo Excel con los resultados del REP A...\n")
                input("\nPresione enter para continuar...\n")
                clear_screen()
                #######################################################################################################################
                # REP B
                clear_screen()
                print("REP B Texto original sin STOPWORDS de los documentos TXTs")


                 # Obtener un índice dentro del rango de la lista
                indice_docu = numero2-1
                # Obtener el elemento correspondiente al índice aleatorio
                elemento = lista_archivos_txt[indice_docu]
                # Imprimir el índice y el elemento elegido
                print("\nEl documento elegido para la prueba es el : {}\n".format(elemento))
                copialistatxts=list(lista_archivos_txt)
                # Eliminar el elemento en el índice de la lista necesario para evitar valores repetidos
                del copialistatxts[indice_docu]



                #Copia de la lista de documentos sin stopwords para trabajar
                aux_documentos_txt=list(docutxtlistlimpios)

                #Se establece los elementos del documento de prueba
                docprueba=aux_documentos_txt[indice_docu]
                del aux_documentos_txt[indice_docu]


                #Se comienza hacer la comparacion
                max_similitud = 0.0
                doc1_max = ""
                similitudes = []

                print("\nResultado de comparaciones...\n")
                for i in range(len(aux_documentos_txt)):
                    similitud = compare_lists(docprueba, aux_documentos_txt[i])
                    if similitud > max_similitud:
                        max_similitud = similitud
                        doc1_max = copialistatxts[i]
                    print("Similitud coseno: {:.2f} entre el {} con el  {}".format(similitud,elemento,copialistatxts[i]))
                    similitudes.append((similitud, copialistatxts[i],' '.join(aux_documentos_txt[i])))
                print("\nLa máxima similitud de {:.2f} se encuentra entre el {} y el {}\n".format(max_similitud,elemento,doc1_max))

                ranking=mostrar_ranking(similitudes)
                
                #Se restaura los valores originales para no alterar en una segunda iteracion
                copialistatxts=list(lista_archivos_txt)
                aux_documentos_txt=list(docutxtlistlimpios)

                generar_excel(ranking,"RepB_SinStopWords_TXTs.xlsx","FFA500")
                modificar_formato_columnas_xlsx("RepB_SinStopWords_TXTs.xlsx")
                print("\nSe genero un archivo Excel con los resultados del REP B...\n")


                input("Presione enter para continuar...")

            elif subopcion == "3":
                clear_screen()
                print("REP A Realizando STEMMING de los documentos PDFs")

                # Obtener un índice dentro del rango de la lista
                indice_docu = numero-1
                # Obtener el elemento correspondiente al índice aleatorio
                elemento = lista_archivos_pdf[indice_docu]
                # Imprimir el índice y el elemento elegido
                print("\nEl documento elegido para la prueba es el : {}\n".format(elemento))

                copialistapdfs=list(lista_archivos_pdf)
                # Eliminar el elemento en el índice de la lista necesario para evitar valores repetidos
                del copialistapdfs[indice_docu]


                #Copia de la lista de documentos sin stopwords para trabajar
                aux_documentos_pdf=list(docupdflistlimpios)

                listdocustemm=[]

                for docu in aux_documentos_pdf:
                    docu=algoritmo_snowball(docu)
                    listdocustemm.append(docu)
                

                #Se establece los elementos del documento de prueba
                docprueba=listdocustemm[indice_docu]
                del listdocustemm[indice_docu]

      
                #Se comienza hacer la comparacion
                max_similitud = 0.0
                doc1_max = ""
                similitudes = []

                

                print("\nResultado de comparaciones...\n")
                for i in range(len(listdocustemm)):
                    similitud = compare_lists(docprueba, listdocustemm[i])
                    if similitud > max_similitud:
                        max_similitud = similitud
                        doc1_max = copialistapdfs[i]
                    print("Similitud coseno: {:.2f} entre el {} con el  {}".format(similitud,elemento,copialistapdfs[i]))
                    similitudes.append((similitud, copialistapdfs[i],' '.join(listdocustemm[i])))
                print("\nLa máxima similitud de {:.2f} se encuentra entre el {} y el {}\n".format(max_similitud,elemento,doc1_max))
                
                ranking=mostrar_ranking(similitudes)

                #Se restaura los valores originales para no alterar en una segunda iteracion
                copialistapdfs=list(lista_archivos_pdf)

                generar_excel(ranking,"RepA_Stemming_PDFs.xlsx","00FF00")
                modificar_formato_columnas_xlsx("RepA_Stemming_PDFs.xlsx")
                print("\nSe genero un archivo Excel con los resultados del REP A...\n")
                input("\nPresione enter para continuar...\n")
                clear_screen()
                #######################################################################################################################
                # REP B
                clear_screen()
                print("REP B Realizando STEMMING de los documentos TXTs")

                # Obtener un índice dentro del rango de la lista
                indice_docu = numero2-1
                # Obtener el elemento correspondiente al índice aleatorio
                elemento = lista_archivos_txt[indice_docu]
                # Imprimir el índice y el elemento elegido
                print("\nEl documento elegido para la prueba es el : {}\n".format(elemento))

                copialistatxts=list(lista_archivos_txt)
                # Eliminar el elemento en el índice de la lista necesario para evitar valores repetidos
                del copialistatxts[indice_docu]


                #Copia de la lista de documentos sin stopwords para trabajar
                aux_documentos_txt=list(docutxtlistlimpios)

                listdocustemm=[]

                for docu in aux_documentos_txt:
                    docu=algoritmo_snowball(docu)
                    listdocustemm.append(docu)
                

                #Se establece los elementos del documento de prueba
                docprueba=listdocustemm[indice_docu]
                del listdocustemm[indice_docu]

      
                #Se comienza hacer la comparacion
                max_similitud = 0.0
                doc1_max = ""
                similitudes = []

                print("\nResultado de comparaciones...\n")
                for i in range(len(listdocustemm)):
                    similitud = compare_lists(docprueba, listdocustemm[i])
                    if similitud > max_similitud:
                        max_similitud = similitud
                        doc1_max = copialistatxts[i]
                    print("Similitud coseno: {:.2f} entre el {} con el  {}".format(similitud,elemento,copialistatxts[i]))
                    similitudes.append((similitud, copialistatxts[i],' '.join(listdocustemm[i])))
                print("\nLa máxima similitud de {:.2f} se encuentra entre el {} y el {}\n".format(max_similitud,elemento,doc1_max))
                
                ranking=mostrar_ranking(similitudes)

                #Se restaura los valores originales para no alterar en una segunda iteracion
                copialistatxts=list(lista_archivos_txt)

                generar_excel(ranking,"RepB_Stemming_TXTs.xlsx","00FF00")
                modificar_formato_columnas_xlsx("RepB_Stemming_TXTs.xlsx")
                print("\nSe genero un archivo Excel con los resultados del REP B...\n")
         

                input("Presione enter para continuar...")


            elif subopcion == "4":
                clear_screen()
                print("REP A Realizando BI-GRAMAS de los documentos PDFs")

                # Obtener un índice dentro del rango de la lista
                indice_docu = numero-1
                # Obtener el elemento correspondiente al índice aleatorio
                elemento = lista_archivos_pdf[indice_docu]
                # Imprimir el índice y el elemento elegido
                print("\nEl documento elegido para la prueba es el : {}\n".format(elemento))

                copialistapdfs=list(lista_archivos_pdf)
                # Eliminar el elemento en el índice de la lista necesario para evitar valores repetidos
                del copialistapdfs[indice_docu]


                #Copia de la lista de documentos sin stopwords para trabajar
                aux_documentos_pdf=list(docupdflistlimpios)

                listdocubigramm=[]

                for docu in aux_documentos_pdf:
                    docu=obtener_bigramas(docu)
                    listdocubigramm.append(docu)
                
                #Se establece los elementos del documento de prueba
                docprueba=listdocubigramm[indice_docu]
                del listdocubigramm[indice_docu]


                #Se comienza hacer la comparacion
                max_similitud = 0.0
                doc1_max = ""
                similitudes=[]

                print("\nResultado de comparaciones...\n")
                for i in range(len(listdocubigramm)):
                    similitud = compare_lists(docprueba, listdocubigramm[i])
                    if similitud > max_similitud:
                        max_similitud = similitud
                        doc1_max = copialistapdfs[i]
                    print("Similitud coseno: {:.2f} entre el {} con el  {}".format(similitud,elemento,copialistapdfs[i]))
                    similitudes.append((similitud, copialistapdfs[i],' '.join(listdocubigramm[i])))
                print("\nLa máxima similitud de {:.2f} se encuentra entre el {} y el {}\n".format(max_similitud,elemento,doc1_max))
                ranking=mostrar_ranking(similitudes)
                
                #Se restaura los valores originales para no alterar en una segunda iteracion
                copialistapdfs=list(lista_archivos_pdf)

                generar_excel(ranking,"RepA_BIGRAMAS_PDFs.xlsx","ADD8E6")
                modificar_formato_columnas_xlsx("RepA_BIGRAMAS_PDFs.xlsx")
                print("\nSe genero un archivo Excel con los resultados del REP A...\n")
                input("\nPresione enter para continuar...\n")
                clear_screen()
                #######################################################################################################################
                # REP B
                clear_screen()
                print("REP B Realizando BI-GRAMAS de los documentos PDFs")

                # Obtener un índice dentro del rango de la lista
                indice_docu = numero2-1
                # Obtener el elemento correspondiente al índice aleatorio
                elemento = lista_archivos_txt[indice_docu]
                # Imprimir el índice y el elemento elegido
                print("\nEl documento elegido para la prueba es el : {}\n".format(elemento))

                copialistatxts=list(lista_archivos_txt)
                # Eliminar el elemento en el índice de la lista necesario para evitar valores repetidos
                del copialistatxts[indice_docu]


                #Copia de la lista de documentos sin stopwords para trabajar
                aux_documentos_txt=list(docutxtlistlimpios)

                listdocubigramm=[]

                for docu in aux_documentos_txt:
                    docu=obtener_bigramas(docu)
                    listdocubigramm.append(docu)
                
                #Se establece los elementos del documento de prueba
                docprueba=listdocubigramm[indice_docu]
                del listdocubigramm[indice_docu]


                #Se comienza hacer la comparacion
                max_similitud = 0.0
                doc1_max = ""
                similitudes=[]

                print("\nResultado de comparaciones...\n")
                for i in range(len(listdocubigramm)):
                    similitud = compare_lists(docprueba, listdocubigramm[i])
                    if similitud > max_similitud:
                        max_similitud = similitud
                        doc1_max = copialistatxts[i]
                    print("Similitud coseno: {:.2f} entre el {} con el  {}".format(similitud,elemento,copialistatxts[i]))
                    similitudes.append((similitud, copialistatxts[i],' '.join(listdocubigramm[i])))
                print("\nLa máxima similitud de {:.2f} se encuentra entre el {} y el {}\n".format(max_similitud,elemento,doc1_max))

                ranking=mostrar_ranking(similitudes)

                #Se restaura los valores originales para no alterar en una segunda iteracion
                copialistapdfs=list(lista_archivos_txt)

                generar_excel(ranking,"RepB_BIGRAMAS_TXTs.xlsx","ADD8E6")
                modificar_formato_columnas_xlsx("RepB_BIGRAMAS_TXTs.xlsx")
                print("\nSe genero un archivo Excel con los resultados del REP B...\n")

                input("Presione enter para continuar...") 
            elif subopcion == "5":
                break
            else:
                print("Opción no válida, por favor intente de nuevo.")
                input("Presione enter para continuar...")
    
    
    elif opcion == "3":
        print("Saliendo del programa...")
        break
    
    else:
        print("Opción no válida, por favor intente de nuevo.")
        input("Presione enter para continuar...")
